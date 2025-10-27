from openpyxl import load_workbook
from openpyxl.styles import Font
import numpy as np


wb = load_workbook("data/Arkusz_25_26.xlsx")
# ws = wb["Arkusz org 2023-2024"]
ws = wb["Arkusz org 2008-2009 1.09.08"]

classes = [
    (ind, c[0], c[1] + "_" + c[2])
    for ind, c in enumerate(zip(*ws.iter_rows(values_only=True, max_row=3)))
    if all(c[i] for i in range(3))
]
print(classes)

subjects = [
    (ind, s.value)
    for ind, s in enumerate(*ws.iter_cols(min_col=2, max_col=2, min_row=2), start=1)
    if s and s.font and s.font.bold
]

teachers = [
    (ind, t.value)
    for ind, t in enumerate(*ws.iter_cols(min_col=2, max_col=2, min_row=4), start=1)
    if t.value and "." in t.value and not t.font.bold
]

# with open("teachers.txt", "w") as file:
#     file.writelines(map(lambda t: t[1] + "\n", teachers))

with open("student_groups.txt", "w") as file:
    file.writelines(map(lambda t: t[2] + "," + t[1] + "\n", classes))


def get_teacher_subject_ind(teacher_row: int, return_name: bool = False) -> int:
    for ind, (s, name) in reversed(list(enumerate(subjects))):
        if s < teacher_row:
            return ind if not return_name else name


# constraint_matrix = np.zeros(
#     shape=(len(teachers), len(classes), len(subjects)), dtype=np.uint8
# )

# rows = list(ws.iter_rows(values_only=True, max_col=28))
# for t_i, (t, teacher) in enumerate(teachers):
#     for c_i, (c, _, class_name) in enumerate(classes):
#         if rows[t][c]:
#             constraint_matrix[t_i, c_i, get_teacher_subject_ind(t)] = np.uint8(
#                 int(rows[t][c])
#             )

# for ind, (_, c, c2) in enumerate(classes):
#     print(ind, c, c2)

# print("---------")
# tmp = {}
# for ind, (_, name) in enumerate(subjects):
#     print(name)
#     tmp[ind] = name
# print(tmp)

# print("---------")
# for ind, (_, t) in enumerate(teachers):
#     print(ind, t)

global_complementary_subjects = [
    (11, 12, 14),  # Języki
    (1, 8),  # Angielski + inf
    (1, 1),  # Ang + Ang
    (10, 10, 4),  # Religia + etyka
    (16, 16),  # WF
]

per_class_complementary_subjects = {
    6: ((15, 3),),  # Psych-prawna  # Biol + WOS
    7: ((11, 18),),  # Niem + mat
    8: ((6, 8), (15, 7)),  # Chem + inf  # Biol + fiz
    9: ((6, 15)),  # >30 osób  # Chem + biol
}

# np.save("data/constraints.npy", constraint_matrix)
# print(constraint_matrix)
# print(constraint_matrix[:, :, 0])
# print(constraint_matrix.shape)
